"""
parse-ig-sales-report
=====================
Azure Function (v1 programming model, HTTP trigger).

Called by a Power Automate flow that monitors igreports@aubreys.group
for emails whose subject matches the pattern NN_Daily (where NN is a
zero-padded location ID, e.g. "05_Daily").  Power Automate extracts the
location ID from the subject and passes both the ID and the base64-
encoded Excel attachment in the JSON body.

Request body (JSON)
-------------------
{
    "location_id":  5,
    "report_data":  "<base64-encoded .xlsx bytes>"
}

Report format
-------------
IG generates a "Closed Checks" report grouped by Check Type then Table
Number. The workbook layout is:

    Row 0:  "Closed Checks"
    Row 2:  "Processed Business Period Starting M/D/YYYY H:MM AM and Ending ..."
    Row 4:  "Grouped by: Check Type, Table Number"    ← integrity check
    Row 9:  "Selected For: Store = ..."
    ...
    "Check Type:<value>"                              ← L1 grouping header
    "    Table Number:<value>"                        ← L2 grouping header (leading spaces)
    "Closed Date/Time"  ...column headers...          ← repeated per Table Number block
    <data rows>
    "Table Number Total" ...                          ← sub-total row (skip)
    "# of transactions: " ...                         ← sub-count row (skip)
    ... (next Table Number block) ...
    "Check Type Total" ...                            ← section total row (skip)
    "# of transactions: " ...                         ← section count row (skip)
    ... (next Check Type block) ...
    "Grand Total" ...                                 ← stop parsing here

The Check Type value and Table Number value from each grouping header
are injected into every data row under that header as "checkType" and
"tableNumber" fields.

Behaviour
---------
1. Decodes and opens the Excel workbook in memory.
2. Validates that the report is grouped by "Check Type, Table Number"
   (integrity check — rejects reports with a different grouping).
3. Parses the "Processed Business Period" header line to get the report
   start/end datetimes.
4. Derives a business date per row using a 4:00 AM cutoff:
       closedDateTime.hour >= 4  ->  businessDate = closedDateTime.date()
       closedDateTime.hour <  4  ->  businessDate = closedDateTime.date() - 1 day
   This correctly handles both single-day and multi-day (back-fill) reports.
5. Checks both ig.hs_sales_landing AND ig.hs_sales for any existing rows
   covering the same locId + business date range.  Dates that already have
   data are skipped (not re-uploaded); clean dates proceed normally.  This
   allows a multi-day back-fill report to be emailed even when some dates
   in the range are already loaded — only the missing dates are imported.
   All skipped dates are logged as warnings and included in the response.
6. Extracts all data rows for clean dates (skipping sub-total, sub-count,
   section-total, and section-count rows), and uploads a JSON array to
   Azure Blob Storage -- one blob per business date.
7. Returns a JSON summary on success, including any skipped dates.

The downstream ADF/Power Automate pipeline reads the blob and bulk-inserts
into ig.hs_sales_landing.

Environment variables required
-------------------------------
SQL_SERVER              Azure SQL server hostname
SQL_DATABASE            Database name
SQL_USER                SQL login username
SQL_PASSWORD            SQL login password
DATALAKE_IG_SALES_DATA  Full base URL for the sales blob folder,
                        e.g. https://aubdlakegen2.blob.core.windows.net/cherokee/ig_sales/data
DATALAKE_SAS            Shared access signature (existing, used across all datalake functions)
"""

import azure.functions as func
import base64
import io
import json
import logging
import os
import re
import pymssql
from datetime import datetime, timedelta, date
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from azure.storage.blob import BlobClient, ContentSettings


# ── Constants ──────────────────────────────────────────────────────────────────

BUSINESS_DAY_START_HOUR = 4

# Expected value in the "Grouped by:" header row — enforced as an integrity check.
EXPECTED_GROUPING = "Check Type, Table Number"

# Row markers — rows whose first cell starts with any of these are skipped or
# signal the end of the data section.
SKIP_ROW_PREFIXES = (
    "Table Number Total",
    "Check Type Total",
    "# of transactions:",
)
STOP_ROW_PREFIX = "Grand Total"

# L1 / L2 grouping header prefixes (after stripping leading whitespace)
CHECK_TYPE_PREFIX  = "Check Type:"
TABLE_NUMBER_PREFIX = "Table Number:"

# The column header sentinel that marks the start of a data block
COL_HEADER_SENTINEL = "Closed Date/Time"


# ── Helpers ────────────────────────────────────────────────────────────────────

def business_date_from_closed(closed_dt, fallback):
    """
    Return the business date for a transaction.
    Transactions before 4:00 AM belong to the previous business day.
    """
    if closed_dt is None:
        return fallback
    if closed_dt.hour < BUSINESS_DAY_START_HOUR:
        return (closed_dt - timedelta(days=1)).date()
    return closed_dt.date()


def safe_decimal(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        return None


def safe_int(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def first_cell(row):
    """Return the first cell value as a stripped string, or '' if row is empty."""
    if not row:
        return ""
    return str(row[0] or "").strip()


# ── Report validation ──────────────────────────────────────────────────────────

def validate_grouping(ws_rows):
    """
    Integrity check: confirm the report is grouped by 'Check Type, Table Number'.
    Looks for a row starting with 'Grouped by:' and verifies the value matches.
    Raises ValueError if the grouping is missing or unexpected.
    """
    for row in ws_rows[:15]:  # The header is always in the first few rows
        cell = first_cell(row)
        if cell.startswith("Grouped by:"):
            grouping_value = cell[len("Grouped by:"):].strip()
            if grouping_value != EXPECTED_GROUPING:
                raise ValueError(
                    f"Unexpected report grouping: '{grouping_value}'. "
                    f"Expected: '{EXPECTED_GROUPING}'. "
                    f"This function only processes reports grouped by Check Type then Table Number."
                )
            return  # Found and valid
    raise ValueError(
        f"Could not find 'Grouped by:' header in the first 15 rows of the report. "
        f"This does not appear to be the expected Closed Checks report format."
    )


def parse_report_period(ws_rows):
    """
    Extract start/end datetimes from the report header.

    Expected header row (index 2):
        "Processed Business Period Starting 5/27/2026 4:00 AM and Ending 5/28/2026 3:59 AM"
    """
    header_text = ""
    for cell in ws_rows[2]:
        if cell is not None:
            header_text = str(cell)
            break

    pattern = (
        r"Starting\s+(\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}\s+[AP]M)"
        r"\s+and\s+Ending\s+(\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}\s+[AP]M)"
    )
    match = re.search(pattern, header_text, re.IGNORECASE)
    if not match:
        raise ValueError(
            f"Could not parse business period from report header. "
            f"Header text found: {header_text!r}"
        )

    fmt = "%m/%d/%Y %I:%M %p"
    start_dt = datetime.strptime(match.group(1).strip(), fmt)
    end_dt   = datetime.strptime(match.group(2).strip(), fmt)
    return start_dt, end_dt


# ── Row classification ─────────────────────────────────────────────────────────

def classify_row(row):
    """
    Classify a worksheet row into one of several categories.

    Returns one of:
        'empty'         — all cells are None / blank
        'stop'          — Grand Total row; caller should halt iteration
        'skip'          — sub-total, sub-count, or section-total row
        'check_type'    — L1 grouping header ("Check Type:<value>")
        'table_number'  — L2 grouping header ("    Table Number:<value>")
        'col_header'    — repeated column header row ("Closed Date/Time ...")
        'data'          — transaction data row
    """
    if all(c is None or str(c).strip() == "" for c in row):
        return "empty"

    cell = first_cell(row)

    if cell.startswith(STOP_ROW_PREFIX):
        return "stop"

    if any(cell.startswith(p) for p in SKIP_ROW_PREFIXES):
        return "skip"

    # Check Type header — no leading whitespace in the raw cell value
    if cell.startswith(CHECK_TYPE_PREFIX):
        return "check_type"

    # Table Number header — raw cell has leading spaces; first_cell strips them
    if cell.startswith(TABLE_NUMBER_PREFIX):
        return "table_number"

    if cell == COL_HEADER_SENTINEL:
        return "col_header"

    return "data"


# ── Report parsing ─────────────────────────────────────────────────────────────

def parse_data_rows(ws_rows, loc_id, report_start_dt, report_end_dt):
    """
    Parse all data rows from the workbook, injecting the current Check Type
    and Table Number values as additional fields on each record.

    Returns a dict keyed by business date so the caller can write one
    blob per business date -- important for multi-day back-fill reports.
    """
    fallback_date    = report_start_dt.date()
    upload_ts        = datetime.utcnow().isoformat()
    rows_by_date     = {}

    current_check_type   = None
    current_table_number = None

    # Skip the preamble rows (title, period, metadata) — scanning starts from
    # row 0 and the classifier handles everything.
    for row in ws_rows:
        kind = classify_row(row)

        if kind == "stop":
            break

        if kind in ("empty", "skip", "col_header"):
            continue

        if kind == "check_type":
            current_check_type   = first_cell(row)[len(CHECK_TYPE_PREFIX):]
            current_table_number = None   # reset — new L1 block
            continue

        if kind == "table_number":
            current_table_number = first_cell(row)[len(TABLE_NUMBER_PREFIX):]
            continue

        # kind == "data"
        # Rows that appear before the first "Check Type:" header are preamble
        # rows (report title, store name, etc.) and should be skipped.
        if current_check_type is None:
            continue

        closed_dt = row[0] if isinstance(row[0], datetime) else None
        biz_date  = business_date_from_closed(closed_dt, fallback_date)

        record = {
            "locId":          loc_id,
            "businessDate":   biz_date.isoformat(),
            "reportStartDt":  report_start_dt.isoformat(),
            "reportEndDt":    report_end_dt.isoformat(),
            "checkType":      current_check_type,
            "tableNumber":    current_table_number,
            "closedDateTime": closed_dt.isoformat() if closed_dt else None,
            "checkNo":        safe_int(row[1]),
            "ctId":           safe_int(row[2]),
            "mpId":           safe_int(row[3]),
            "serverId":       safe_int(row[4]),
            "cashierId":      safe_int(row[5]),
            "grossRevenue":   safe_decimal(row[6]),
            "discount":       safe_decimal(row[7]),
            "tax":            safe_decimal(row[8]),
            "gratSvcChg":     safe_decimal(row[9]),
            "tip":            safe_decimal(row[10]),
            "roundedAmount":  safe_decimal(row[11]),
            "checkTotal":     safe_decimal(row[12]),
            "tenderId":       safe_int(row[13]),
            "tender":         safe_decimal(row[14]),
            "changeAmt":      safe_decimal(row[15]),
            "breakage":       safe_decimal(row[16]),
            "netTender":      safe_decimal(row[17]),
            "tpl_upload":     upload_ts,
        }

        rows_by_date.setdefault(biz_date, []).append(record)

    return rows_by_date


# ── Database ───────────────────────────────────────────────────────────────────

def get_db_connection():
    return pymssql.connect(
        server=os.environ["SQL_SERVER"],
        database=os.environ["SQL_DATABASE"],
        user=os.environ["SQL_USER"],
        password=os.environ["SQL_PASSWORD"],
    )


def check_existing_data(loc_id, business_dates):
    """
    Return the subset of business_dates that already have data in either
    ig.hs_sales_landing or ig.hs_sales.
    """
    if not business_dates:
        return []

    placeholders = ", ".join(["%s"] * len(business_dates))
    date_strs    = [d.isoformat() for d in business_dates]

    query = f"""
        SELECT DISTINCT businessDate
        FROM (
            SELECT businessDate FROM ig.hs_sales_landing
            WHERE locId = %s AND businessDate IN ({placeholders})
            UNION ALL
            SELECT businessDate FROM ig.hs_sales
            WHERE locId = %s AND businessDate IN ({placeholders})
        ) combined
    """
    params = [loc_id] + date_strs + [loc_id] + date_strs

    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(query, params)
        rows = cursor.fetchall()

    return [row[0] for row in rows]


# ── Blob upload ────────────────────────────────────────────────────────────────

def upload_to_blob(loc_id, biz_date, records):
    """Upload a JSON array to blob storage using SAS and return the blob path."""
    base_url  = os.environ["DATALAKE_IG_SALES_DATA"].rstrip("/")
    sas_token = os.environ["DATALAKE_SAS"].lstrip("?")
    timestamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    blob_name = (
        f"loc{loc_id:02d}/{biz_date.isoformat()}/"
        f"hs_sales_{loc_id:02d}_{biz_date.isoformat()}_{timestamp}.json"
    )

    blob_client = BlobClient.from_blob_url(f"{base_url}/{blob_name}?{sas_token}")
    blob_client.upload_blob(
        json.dumps(records, default=str),
        overwrite=True,
        content_settings=ContentSettings(content_type="application/json"),
    )

    # Return path without SAS for logging/response
    return f"{base_url}/{blob_name}"


# ── Entry point ────────────────────────────────────────────────────────────────

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("parse-ig-sales-report: triggered")

    # Parse request body
    try:
        body = req.get_json()
    except ValueError:
        return func.HttpResponse("Request body must be valid JSON.", status_code=400)

    location_id     = body.get("location_id")
    report_data_b64 = body.get("report_data")

    if location_id is None or not report_data_b64:
        return func.HttpResponse(
            "Request body must include 'location_id' (int) and 'report_data' (base64 string).",
            status_code=400,
        )

    try:
        location_id = int(location_id)
        excel_bytes = base64.b64decode(report_data_b64)
    except Exception as e:
        return func.HttpResponse(f"Invalid input: {e}", status_code=400)

    # Open workbook
    try:
        wb      = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        ws      = wb.active
        ws_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        logging.error(f"parse-ig-sales-report: workbook parse failed: {e}")
        return func.HttpResponse(f"Failed to open Excel file: {e}", status_code=422)

    # Integrity check: confirm report is grouped by Check Type, Table Number
    try:
        validate_grouping(ws_rows)
    except ValueError as e:
        logging.error(f"parse-ig-sales-report: grouping validation failed: {e}")
        return func.HttpResponse(str(e), status_code=422)

    # Parse report period from header
    try:
        report_start_dt, report_end_dt = parse_report_period(ws_rows)
    except ValueError as e:
        logging.error(f"parse-ig-sales-report: header parse failed: {e}")
        return func.HttpResponse(str(e), status_code=422)

    logging.info(
        f"parse-ig-sales-report: locId={location_id}  "
        f"period={report_start_dt} -> {report_end_dt}"
    )

    # Parse data rows
    try:
        rows_by_date = parse_data_rows(ws_rows, location_id, report_start_dt, report_end_dt)
    except Exception as e:
        logging.error(f"parse-ig-sales-report: data parse failed: {e}")
        return func.HttpResponse(f"Failed to parse report rows: {e}", status_code=422)

    if not rows_by_date:
        return func.HttpResponse(
            json.dumps({"status": "skipped", "reason": "No data rows found in report."}),
            mimetype="application/json",
            status_code=200,
        )

    all_business_dates = sorted(rows_by_date.keys())

    # Duplicate check against both landing and archive tables.
    # Conflicting dates are skipped individually; clean dates proceed.
    # This allows partial back-fills when some dates in a multi-day
    # report are already loaded.
    try:
        conflicting_dates = check_existing_data(location_id, all_business_dates)
    except Exception as e:
        logging.error(f"parse-ig-sales-report: DB duplicate check failed: {e}")
        return func.HttpResponse(f"Database check failed: {e}", status_code=500)

    skipped_dates = []
    if conflicting_dates:
        conflict_set  = {
            d.isoformat() if isinstance(d, date) else str(d)
            for d in conflicting_dates
        }
        skipped_dates = sorted(conflict_set)
        for skipped in skipped_dates:
            logging.warning(
                f"parse-ig-sales-report: locId={location_id} date={skipped} "
                f"already exists in ig.hs_sales_landing or ig.hs_sales — skipped."
            )
        # Remove conflicting dates from the upload set
        rows_by_date = {
            d: r for d, r in rows_by_date.items()
            if d.isoformat() not in conflict_set
        }

    if not rows_by_date:
        msg = (
            f"All {len(skipped_dates)} date(s) in this report already have data "
            f"for locId={location_id}. Nothing to upload."
        )
        logging.warning(f"parse-ig-sales-report: {msg}")
        return func.HttpResponse(
            json.dumps({
                "status":       "skipped",
                "reason":       msg,
                "skippedDates": skipped_dates,
            }),
            mimetype="application/json",
            status_code=200,
        )

    # Upload one blob per business date
    results = []
    for biz_date, records in rows_by_date.items():
        try:
            blob_path = upload_to_blob(location_id, biz_date, records)
            logging.info(
                f"parse-ig-sales-report: uploaded {len(records)} rows -> {blob_path}"
            )
            results.append({
                "businessDate": biz_date.isoformat(),
                "rowCount":     len(records),
                "blobPath":     blob_path,
            })
        except Exception as e:
            logging.error(f"parse-ig-sales-report: blob upload failed for {biz_date}: {e}")
            return func.HttpResponse(
                f"Blob upload failed for {biz_date}: {e}", status_code=500
            )

    return func.HttpResponse(
        json.dumps({
            "status":       "success",
            "locId":        location_id,
            "reportStart":  report_start_dt.isoformat(),
            "reportEnd":    report_end_dt.isoformat(),
            "uploads":      results,
            "skippedDates": skipped_dates,
        }),
        mimetype="application/json",
        status_code=200,
    )